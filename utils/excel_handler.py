"""
utils/excel_handler.py — Read document requirements and write RAG verdicts
                          into the HR Candidate Verification Excel.

Excel structure (fixed):
  Row 1  : Title banner (merged)
  Row 2  : Candidate info fields (name, position, interview date, HR manager)
  Row 3  : Spacer
  Row 4  : Completion rate formula banner
  Row 5  : Spacer
  Row 6  : Column headers  → used as pandas header row (index 5, 0-based)
  Row 7+ : Document rows   → data starts here

Column mapping (1-based):
  A (1)  #                    — row number, ignored
  B (2)  Document (FR)        → used as the RAG search query
  C (3)  Document (EN)        → stored for context
  D (4)  Obligatoire?         → determines fallback status when not found
  E (5)  Statut               ← RAG writes: ✅ Soumis / ❌ Manquant / N/A
  F (6)  Vérifié par          ← RAG writes: "AI RAG System"
  G (7)  Date de vérification ← RAG writes: today's date
  H (8)  Date d'expiration    — left untouched (HR fills manually)
  I (9)  Commentaires         ← RAG writes: justification + source reference
  J (10) Alerte               — contains Excel formulas, never overwritten
"""
import logging
from datetime import date
from typing import Dict, List

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

logger = logging.getLogger(__name__)

# ─── Layout constants ─────────────────────────────────────────────────────────
DATA_START   = 7   # Excel row where document entries begin
DATA_END     = 26  # Excel row where document entries end (20 docs)

# Column indices (1-based, matches openpyxl convention)
COL_DOC_FR   = 2   # B — Document name in French  (RAG query)
COL_DOC_EN   = 3   # C — Document name in English (context)
COL_REQUIRED = 4   # D — Required? (Oui / Non / Si applicable)
COL_STATUS   = 5   # E — Statut  <- written by RAG
COL_VERIF_BY = 6   # F — Verifie par  <- written by RAG
COL_DATE     = 7   # G — Date de verification  <- written by RAG
COL_NOTES    = 9   # I — Commentaires  <- written by RAG
# Column J (10) contains Excel formulas — never touched

VERIFIER_LABEL = "AI RAG System"

# ─── Status values (must match the Excel dropdown list) ───────────────────────
STATUS_FOUND   = "✅ Soumis"
STATUS_MISSING = "❌ Manquant"
STATUS_NA      = "N/A"

# ─── Styles ───────────────────────────────────────────────────────────────────
_thin   = Side(style="thin", color="AAAAAA")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

_GREEN_BG = PatternFill("solid", start_color="C6EFCE")
_RED_BG   = PatternFill("solid", start_color="FFC7CE")
_GREY_BG  = PatternFill("solid", start_color="F2F2F2")
_GREEN_FG = "276221"
_RED_FG   = "9C0006"
_BODY_FONT = Font(name="Arial", size=9)


def _status_fill(satisfied: bool, is_required: bool) -> PatternFill:
    if not satisfied and not is_required:
        return _GREY_BG
    return _GREEN_BG if satisfied else _RED_BG


def _status_font(satisfied: bool) -> Font:
    color = _GREEN_FG if satisfied else _RED_FG
    return Font(name="Arial", bold=True, size=9, color=color)


# ─── Reading ──────────────────────────────────────────────────────────────────

def read_document_checklist(filepath: str) -> List[Dict]:
    """
    Read the HR candidate verification Excel and return one dict per document row.

    Each dict contains:
        doc_fr     : str  — French document name  (used as RAG query)
        doc_en     : str  — English document name (extra context for LLM)
        required   : str  — Raw "Obligatoire?" cell value
        is_required: bool — True when "Oui" appears in the required field

    Rows with no document name are silently skipped.
    """
    # header=5 → row index 5 (0-based) = Excel row 6
    df = pd.read_excel(filepath, sheet_name=0, header=5)

    rows = []
    for _, row in df.iterrows():
        doc_fr   = str(row.get("Document (FR)",           "")).strip()
        doc_en   = str(row.get("Document (EN)",           "")).strip()
        required = str(row.get("Obligatoire?\nRequired?", "")).strip()

        # Skip empty / NaN rows
        if not doc_fr or doc_fr.lower() == "nan":
            continue

        rows.append({
            "doc_fr":       doc_fr,
            "doc_en":       doc_en,
            "required":     required,
            # "Si applicable" is treated as optional — HR decides manually
            "is_required":  "oui" in required.lower(),
        })

    logger.info(f"Read {len(rows)} document requirement(s) from {filepath}")
    return rows


# ─── Writing ──────────────────────────────────────────────────────────────────

def write_results(input_path: str, output_path: str, results: List[Dict]) -> None:
    """
    Write RAG verdicts back into the existing HR verification Excel.

    Fills columns E (Statut), F (Verifie par), G (Date), I (Commentaires)
    for each document row. Column J (Alerte) is intentionally left untouched
    because it already contains Excel formulas that auto-update from column E.

    Args:
        input_path : Original HR Excel file (used as the base template).
        output_path: Where the enriched copy is saved.
        results    : List of {"doc_fr", "is_required", "verdict"} dicts
                     in the same order as the document rows (row 7, 8, ...).
    """
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    today = date.today().strftime("%d/%m/%Y")

    for idx, result in enumerate(results):
        excel_row = DATA_START + idx   # row 7, 8, 9, ...
        if excel_row > DATA_END:
            logger.warning(f"More results than expected rows — stopping at row {DATA_END}")
            break

        verdict     = result.get("verdict", {})
        satisfied   = verdict.get("satisfied", False)
        is_required = result.get("is_required", True)

        # ── Determine status value ────────────────────────────────────────────
        if satisfied:
            status = STATUS_FOUND
        elif not is_required:
            # Optional doc not found → N/A (HR will decide)
            status = STATUS_NA
        else:
            status = STATUS_MISSING

        # ── Build the notes string ─────────────────────────────────────────────
        justification = verdict.get("justification", "")
        source_file   = verdict.get("source_file",   "N/A")
        source_page   = verdict.get("source_page",   0)
        confidence    = verdict.get("confidence",    "").upper()

        source_ref = (
            f"[{source_file}, p.{source_page}]"
            if source_file != "N/A"
            else "[No matching document found]"
        )
        notes = f"[{confidence}] {justification} {source_ref}"

        # ── Write to cells ────────────────────────────────────────────────────
        _write_cell(ws, excel_row, COL_STATUS,   status,          satisfied, is_required)
        _write_cell(ws, excel_row, COL_VERIF_BY, VERIFIER_LABEL,  satisfied, is_required)
        _write_cell(ws, excel_row, COL_DATE,     today,           satisfied, is_required)
        _write_cell(ws, excel_row, COL_NOTES,    notes,           satisfied, is_required, wrap=True)

    wb.save(output_path)
    logger.info(f"Results written to {output_path}")


def _write_cell(ws, row: int, col: int, value: str,
                satisfied: bool, is_required: bool, wrap: bool = False) -> None:
    """Write a value into a cell with appropriate color and border styling."""
    cell = ws.cell(row=row, column=col)
    cell.value     = value
    cell.fill      = _status_fill(satisfied, is_required)
    cell.font      = _status_font(satisfied) if col == COL_STATUS else _BODY_FONT
    cell.border    = _border
    cell.alignment = Alignment(
        horizontal="left" if wrap else "center",
        vertical="center",
        wrap_text=wrap,
    )
