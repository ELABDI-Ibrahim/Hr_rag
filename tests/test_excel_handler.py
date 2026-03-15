"""
tests/test_excel_handler.py — Unit tests for utils/excel_handler.py

Tests:
  - read_document_checklist: correct parsing of the HR Excel structure
  - write_results: cells are written with correct values and an output file is created

NOTE: The tests use a *minimal* openpyxl workbook that mirrors just the columns
the handler needs, so no real HR Excel file is required.

Run:
    pytest tests/test_excel_handler.py -v
"""
from pathlib import Path

import openpyxl
import pytest

from utils.excel_handler import (
    COL_DATE,
    COL_NOTES,
    COL_STATUS,
    COL_VERIF_BY,
    DATA_START,
    STATUS_FOUND,
    STATUS_MISSING,
    STATUS_NA,
    write_results,
)


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _build_minimal_excel(tmp_path: Path, rows: list[dict]) -> Path:
    """
    Build a minimal Excel workbook that mirrors the structure expected by
    read_document_checklist and write_results:
      - Rows 1–5: spacers (blank)
      - Row 6: header row with column names
      - Row 7+: data rows
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    # Rows 1-5 are spacers (write blank so openpyxl doesn't skip them)
    for r in range(1, 6):
        ws.cell(row=r, column=1, value="")

    # Row 6: headers (match exactly what pandas will read)
    headers = [
        "#",
        "Document (FR)",
        "Document (EN)",
        "Obligatoire?\nRequired?",
        "Statut",
        "Vérifié par",
        "Date de vérification",
        "Date d'expiration",
        "Commentaires",
        "Alerte",
    ]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=6, column=col_idx, value=header)

    # Row 7+: data rows
    for i, row in enumerate(rows):
        excel_row = DATA_START + i
        ws.cell(row=excel_row, column=1, value=i + 1)
        ws.cell(row=excel_row, column=2, value=row.get("doc_fr", ""))
        ws.cell(row=excel_row, column=3, value=row.get("doc_en", ""))
        ws.cell(row=excel_row, column=4, value=row.get("required", "Oui"))

    path = tmp_path / "test_hr.xlsx"
    wb.save(str(path))
    return path


# ─── read_document_checklist ──────────────────────────────────────────────────

class TestReadDocumentChecklist:
    def test_reads_correct_number_of_rows(self, tmp_path):
        from utils.excel_handler import read_document_checklist

        xls = _build_minimal_excel(tmp_path, [
            {"doc_fr": "Curriculum Vitae",    "doc_en": "Resume",    "required": "Oui"},
            {"doc_fr": "Pièce d'identité",    "doc_en": "ID Card",   "required": "Oui"},
            {"doc_fr": "Attestation de stage", "doc_en": "Internship cert", "required": "Non"},
        ])
        rows = read_document_checklist(str(xls))
        assert len(rows) == 3

    def test_is_required_true_for_oui(self, tmp_path):
        from utils.excel_handler import read_document_checklist

        xls = _build_minimal_excel(tmp_path, [
            {"doc_fr": "CV", "doc_en": "Resume", "required": "Oui"},
        ])
        rows = read_document_checklist(str(xls))
        assert rows[0]["is_required"] is True

    def test_is_required_false_for_non(self, tmp_path):
        from utils.excel_handler import read_document_checklist

        xls = _build_minimal_excel(tmp_path, [
            {"doc_fr": "Lettre de référence", "doc_en": "Reference letter", "required": "Non"},
        ])
        rows = read_document_checklist(str(xls))
        assert rows[0]["is_required"] is False

    def test_empty_doc_fr_rows_are_skipped(self, tmp_path):
        from utils.excel_handler import read_document_checklist

        xls = _build_minimal_excel(tmp_path, [
            {"doc_fr": "CV",  "doc_en": "Resume", "required": "Oui"},
            {"doc_fr": "",    "doc_en": "",        "required": ""},
        ])
        rows = read_document_checklist(str(xls))
        assert len(rows) == 1


# ─── write_results ────────────────────────────────────────────────────────────

class TestWriteResults:
    def _make_result(self, satisfied: bool, is_required: bool = True):
        verdict = {
            "satisfied":     satisfied,
            "confidence":    "high" if satisfied else "low",
            "justification": "Test justification.",
            "source_file":   "cv.pdf" if satisfied else "N/A",
            "source_page":   1 if satisfied else 0,
        }
        return {"doc_fr": "CV", "is_required": is_required, "verdict": verdict}

    def test_output_file_created(self, tmp_path):
        input_xls  = _build_minimal_excel(tmp_path, [{"doc_fr": "CV", "doc_en": "Resume", "required": "Oui"}])
        output_xls = tmp_path / "output.xlsx"
        write_results(str(input_xls), str(output_xls), [self._make_result(True)])
        assert output_xls.exists()

    def test_found_status_written(self, tmp_path):
        input_xls  = _build_minimal_excel(tmp_path, [{"doc_fr": "CV", "doc_en": "Resume", "required": "Oui"}])
        output_xls = tmp_path / "output_found.xlsx"

        write_results(str(input_xls), str(output_xls), [self._make_result(True)])

        wb  = openpyxl.load_workbook(str(output_xls))
        ws  = wb.active
        val = ws.cell(row=DATA_START, column=COL_STATUS).value
        assert val == STATUS_FOUND

    def test_missing_required_status_written(self, tmp_path):
        input_xls  = _build_minimal_excel(tmp_path, [{"doc_fr": "CV", "doc_en": "Resume", "required": "Oui"}])
        output_xls = tmp_path / "output_missing.xlsx"

        write_results(str(input_xls), str(output_xls), [self._make_result(False, is_required=True)])

        wb  = openpyxl.load_workbook(str(output_xls))
        ws  = wb.active
        val = ws.cell(row=DATA_START, column=COL_STATUS).value
        assert val == STATUS_MISSING

    def test_optional_not_found_writes_na(self, tmp_path):
        input_xls  = _build_minimal_excel(tmp_path, [{"doc_fr": "Lettre", "doc_en": "Letter", "required": "Non"}])
        output_xls = tmp_path / "output_na.xlsx"

        write_results(str(input_xls), str(output_xls), [self._make_result(False, is_required=False)])

        wb  = openpyxl.load_workbook(str(output_xls))
        ws  = wb.active
        val = ws.cell(row=DATA_START, column=COL_STATUS).value
        assert val == STATUS_NA

    def test_verifier_label_written(self, tmp_path):
        from utils.excel_handler import VERIFIER_LABEL

        input_xls  = _build_minimal_excel(tmp_path, [{"doc_fr": "CV", "doc_en": "Resume", "required": "Oui"}])
        output_xls = tmp_path / "output_verifier.xlsx"

        write_results(str(input_xls), str(output_xls), [self._make_result(True)])

        wb  = openpyxl.load_workbook(str(output_xls))
        ws  = wb.active
        val = ws.cell(row=DATA_START, column=COL_VERIF_BY).value
        assert val == VERIFIER_LABEL
